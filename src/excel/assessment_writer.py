import openpyxl
from datetime import datetime
from pathlib import Path
import shutil

class AssessmentWriter:
    def __init__(self, template_path='templates/アセスメントシート原本.xlsx'):
        self.template_path = Path(template_path)
    
    def create_assessment_file(self, interview_data, assessment_data, output_path):
        """アセスメントシートExcelファイルを作成"""
        
        if not self.template_path.exists():
            raise FileNotFoundError(
                f"テンプレートファイルが見つかりません: {self.template_path}\n"
                "templates/ディレクトリにアセスメントシート原本.xlsxを配置してください"
            )
        
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # === 重要：テンプレートをコピーしてから開く ===
        print(f"📋 テンプレートをコピー: {self.template_path} → {output_path}")
        shutil.copy2(str(self.template_path), str(output_path))
        
        try:
            # 書式を保持しながら読み込み
            wb = openpyxl.load_workbook(
                str(output_path),
                data_only=False
            )
            
            # シート名を確認
            sheet_names = wb.sheetnames
            print(f"✅ 利用可能なシート名: {sheet_names}")
            
            # アセスメントシートを探す
            target_sheet = None
            for sheet_name in sheet_names:
                if 'アセスメント' in sheet_name or 'assessment' in sheet_name.lower():
                    target_sheet = sheet_name
                    break
            
            if not target_sheet:
                target_sheet = sheet_names[0]
                print(f"⚠️ 'アセスメントシート'が見つかりません。'{target_sheet}'を使用します")
            
            ws = wb[target_sheet]
            print(f"📝 シート '{target_sheet}' にデータを入力します")
            
        except Exception as e:
            raise ValueError(f"テンプレートファイルの読み込みエラー: {str(e)}")
        
        # === データの書き込み（書式を保持しながら） ===
        
        # 基本情報（Row 3-5）
        self._write_cell_value(ws, 'C3', self._generate_support_number(interview_data['面談実施日']))
        self._write_cell_value(ws, 'G3', interview_data.get('担当支援員', ''))
        
        if isinstance(interview_data['面談実施日'], datetime):
            self._write_cell_value(ws, 'H3', interview_data['面談実施日'].strftime('%m月%d日'))
        else:
            self._write_cell_value(ws, 'H3', interview_data['面談実施日'])
        
        self._write_cell_value(ws, 'C4', interview_data.get('保護者氏名', ''))
        self._write_cell_value(ws, 'G4', interview_data.get('児童氏名', ''))
        self._write_cell_value(ws, 'O4', interview_data.get('性別', ''))
        
        self._write_cell_value(ws, 'C5', interview_data.get('学校名', ''))
        self._write_cell_value(ws, 'I5', str(interview_data.get('学年', '')))
        self._write_cell_value(ws, 'N5', interview_data.get('ひとり親世帯', '該当しない'))
        
        # 課題チェックリスト（各セルに個別入力）
        self._fill_household_issues(ws, assessment_data)
        
        # 希望する進路（B18セル）
        future_path = assessment_data.get('future_path', {})
        if isinstance(interview_data['面談実施日'], datetime):
            confirm_date = interview_data['面談実施日'].strftime('%Y/%m/%d')
        else:
            confirm_date = str(interview_data['面談実施日'])
        
        checkbox_進学 = "■" if future_path.get('type') == "進学" else "□"
        checkbox_就職 = "■" if future_path.get('type') == "就職" else "□"
        
        path_text = f"""確認日　{confirm_date}
{checkbox_進学}進学　　{checkbox_就職}就職
（具体的内容）
・{future_path.get('detail', '')}"""
        
        self._write_cell_value(ws, 'B18', path_text)
        
        # 支援計画（短期目標）Row 29-30
        if 'short_term_plan' in assessment_data:
            self._fill_support_plan(ws, assessment_data['short_term_plan'], start_row=29)
        
        # 本事業における達成目標 Row 35-36
        if 'long_term_plan' in assessment_data:
            self._fill_support_plan(ws, assessment_data['long_term_plan'], start_row=35)
        
        # 保存
        wb.save(str(output_path))
        print(f"✅ アセスメントシートを作成: {output_path}")
        
        return str(output_path)
    
    def _write_cell_value(self, ws, cell_ref, value):
        """
        セルに値を書き込む（既存の書式を保持）
        
        重要：ws[cell_ref].value = value だと書式が崩れる可能性があるため、
        この専用メソッドを使う
        """
        if not value:
            return
        
        cell = ws[cell_ref]
        
        # 既存の書式情報を保存
        original_font = cell.font.copy() if cell.font else None
        original_alignment = cell.alignment.copy() if cell.alignment else None
        original_fill = cell.fill.copy() if cell.fill else None
        original_border = cell.border.copy() if cell.border else None
        original_number_format = cell.number_format
        
        # 値を設定
        cell.value = value
        
        # 書式を復元（値を設定すると消える場合があるため）
        if original_font:
            cell.font = original_font
        if original_alignment:
            cell.alignment = original_alignment
        if original_fill:
            cell.fill = original_fill
        if original_border:
            cell.border = original_border
        if original_number_format:
            cell.number_format = original_number_format
    
    def _generate_support_number(self, date):
        """支援番号を自動生成"""
        if isinstance(date, datetime):
            return date.strftime('%Y%m%d')
        return str(date)
    
    def _format_issues_cell(self, issues_data):
        """
        課題チェックリストのセル（B11）を整形
        
        テンプレートのB11は大きな結合セルなので、
        ここに全ての課題をまとめて書き込む
        """
        lines = []
        
        # 課題の順番（テンプレートに合わせる）
        issue_order = [
            "不登校",
            "引きこもり", 
            "生活リズム",
            "生活習慣",
            "学習の遅れ・低学力",
            "学習習慣・環境",
            "発達特性or発達課題",
            "対人緊張の高さ",
            "コミュニケーションに苦手意識",
            "家庭環境",
            "虐待",
            "他の世帯員の問題",
            "その他"
        ]
        
        for key in issue_order:
            if key in issues_data:
                value = issues_data[key]
                checkbox = "■" if value.get('該当', False) else "□"
                detail = value.get('詳細', '')
                
                if key == "その他" and value.get('該当'):
                    lines.append(f"{checkbox} {key}")
                    if detail:
                        lines.append(f"・{detail}")
                else:
                    if detail and detail not in ["該当なし", "特に問題なし", "不明", ""]:
                        lines.append(f"{checkbox} {key}（{detail}）")
                    else:
                        lines.append(f"{checkbox} {key}")
        
        return "\n".join(lines)
    
    def _fill_support_plan(self, ws, plan_data, start_row):
        """支援計画の表を埋める"""
        
        # テンプレートのセル位置を確認（画像1-4を参考）
        # Row 28: ヘッダー（課題、現状、ニーズ、目標、具体的な方法）
        # Row 29-30: 短期目標のデータ
        # Row 35-36: 長期目標のデータ
        
        self._write_cell_value(ws, f'B{start_row}', plan_data.get('現状', ''))
        self._write_cell_value(ws, f'E{start_row}', plan_data.get('ニーズ_本人', ''))
        self._write_cell_value(ws, f'E{start_row + 1}', plan_data.get('ニーズ_保護者', ''))
        self._write_cell_value(ws, f'I{start_row}', plan_data.get('目標', ''))
        self._write_cell_value(ws, f'M{start_row}', plan_data.get('方法', ''))
    
    def _fill_household_issues(self, ws, assessment_data):
        """世帯の具体的な課題を入力"""
        issues_data = assessment_data.get('issues', {})
        
        # 各課題項目のチェックボックスと詳細を設定
        issue_mappings = {
            '不登校': {'row': 11, 'col': 'B'},
            '引きこもり': {'row': 11, 'col': 'I'},
            '生活リズム': {'row': 12, 'col': 'B'},
            '生活習慣': {'row': 12, 'col': 'I'},
            '学習の遅れ・低学力': {'row': 13, 'col': 'B'},
            '学習習慣・環境': {'row': 13, 'col': 'I'},
            '発達特性or発達課題': {'row': 14, 'col': 'B'},
            '対人緊張の高さ': {'row': 14, 'col': 'I'},
            'コミュニケーションに苦手意識': {'row': 15, 'col': 'B'},
            '家庭環境': {'row': 15, 'col': 'I'},
            '虐待': {'row': 16, 'col': 'B'},
            'その他': {'row': 16, 'col': 'I'}
        }
        
        for issue_name, mapping in issue_mappings.items():
            if issue_name in issues_data:
                issue_data = issues_data[issue_name]
                checkbox = "■" if issue_data.get('該当', False) else "□"
                detail = issue_data.get('詳細', '')
                
                # 既存のセル内容を取得
                cell_address = f"{mapping['col']}{mapping['row']}"
                existing_value = ws[cell_address].value or ""
                
                # テンプレートのフォーマットを保持しながら、チェックボックスと詳細のみ更新
                if existing_value and existing_value.strip():
                    # 既存内容がある場合は、チェックボックスと詳細のみ更新
                    existing_text = str(existing_value)
                    
                    # チェックボックス部分のみを更新（既存の詳細部分は保持）
                    if existing_text.startswith('□') or existing_text.startswith('■'):
                        # 既存のチェックボックスを新しいものに置換
                        # 既存の詳細部分を抽出
                        if '(' in existing_text and ')' in existing_text:
                            # 既存の詳細部分を抽出
                            start_idx = existing_text.find('(')
                            end_idx = existing_text.rfind(')')
                            if start_idx != -1 and end_idx != -1:
                                existing_detail = existing_text[start_idx:end_idx+1]
                                # チェックボックスのみを更新し、既存の詳細部分は保持
                                new_value = f"{checkbox}{issue_name}{existing_detail}"
                            else:
                                # 詳細部分がない場合は、新しい詳細を追加
                                if detail and detail.strip():
                                    new_value = f"{checkbox}{issue_name}({detail})"
                                else:
                                    new_value = f"{checkbox}{issue_name}"
                        else:
                            # 詳細部分がない場合は、新しい詳細を追加
                            if detail and detail.strip():
                                new_value = f"{checkbox}{issue_name}({detail})"
                            else:
                                new_value = f"{checkbox}{issue_name}"
                    else:
                        # チェックボックスがない場合は、新しい内容を設定
                        if detail and detail.strip():
                            new_value = f"{checkbox}{issue_name}({detail})"
                        else:
                            new_value = f"{checkbox}{issue_name}"
                    
                    # セルに新しい値を設定
                    self._write_cell_value(ws, cell_address, new_value)
                else:
                    # 既存内容がない場合は、新しい内容を設定
                    if detail and detail.strip():
                        new_value = f"{checkbox}{issue_name}({detail})"
                    else:
                        new_value = f"{checkbox}{issue_name}"
                    
                    self._write_cell_value(ws, cell_address, new_value)