#!/usr/bin/env python3
"""
アセスメントシート生成スクリプト（Python版）
- テンプレートのフォーマットを完全に保持
- データを埋め込み
- パスワード保護機能付き
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border
import copy
import os
import shutil
from datetime import datetime


def set_cell_value_preserve_format(worksheet, cell_ref, value):
    """
    セルに値を設定し、既存のフォーマットを完全に保持
    
    Args:
        worksheet: ワークシート
        cell_ref: セル参照（例: 'D3'）
        value: 設定する値
    """
    if value is None:
        return
    
    cell = worksheet[cell_ref]
    
    # 既存のスタイルを保存（deepcopy）
    original_font = copy.copy(cell.font)
    original_fill = copy.copy(cell.fill)
    original_border = copy.copy(cell.border)
    original_alignment = copy.copy(cell.alignment)
    original_number_format = cell.number_format
    original_protection = copy.copy(cell.protection)
    
    # 値を設定
    cell.value = value
    
    # スタイルを復元
    cell.font = original_font
    cell.fill = original_fill
    cell.border = original_border
    cell.alignment = original_alignment
    cell.number_format = original_number_format
    cell.protection = original_protection


def format_issues_text(issues):
    """
    課題チェックリストをテキスト化
    
    Args:
        issues: 課題の辞書
        
    Returns:
        フォーマットされたテキスト
    """
    lines = []
    
    issue_order = [
        "不登校",
        "引きこもり", 
        "生活リズム",
        "生活習慣",
        "学習の遅れ・低学力",
        "学習習慣・環境",
        "発達特性or発達課題",
        "対人緊張の高さ",
        "コミュニケーションに苦手意識",
        "家庭環境",
        "虐待",
        "他の世帯員の問題",
        "その他"
    ]
    
    for issue_name in issue_order:
        if issue_name in issues and issues[issue_name]:
            issue = issues[issue_name]
            checkbox = '■' if issue.get('checked') else '□'
            detail = issue.get('detail', '')
            
            if issue_name == "その他" and issue.get('checked'):
                lines.append(f'{checkbox} {issue_name}')
                if detail:
                    lines.append(f'・{detail}')
            else:
                if detail and detail not in ['該当なし', '特に問題なし', '不明', '']:
                    lines.append(f'{checkbox} {issue_name}（{detail}）')
                else:
                    lines.append(f'{checkbox} {issue_name}')
    
    return '\n'.join(lines)


def generate_assessment_sheet(template_path, output_path, data, password=None):
    """
    アセスメントシートを生成
    
    Args:
        template_path: テンプレートファイルのパス
        output_path: 出力ファイルのパス
        data: 面談データの辞書
        password: パスワード（Noneの場合は保護なし）
        
    Returns:
        dict: 結果（success, outputPath）
    """
    print('=' * 70)
    print('アセスメントシート生成')
    print('=' * 70)
    print()
    
    try:
        # テンプレートの存在確認
        if not os.path.exists(template_path):
            raise FileNotFoundError(f'テンプレートが見つかりません: {template_path}')
        
        print(f'📋 テンプレートを読み込み: {template_path}')
        
        # テンプレートを読み込み
        wb = load_workbook(template_path)
        
        # シートを取得
        ws = wb['ｱｾｽﾒﾝﾄｼｰﾄ']
        print(f'✓ シート取得: {ws.title}')
        print(f'  サイズ: {ws.max_row}行 × {ws.max_column}列')
        print(f'  結合セル数: {len(ws.merged_cells.ranges)}')
        print()
        
        # === 基本情報を書き込み ===
        print('📝 データを書き込み中...')
        
        set_cell_value_preserve_format(ws, 'D3', data.get('supportNumber'))      # 支援番号
        set_cell_value_preserve_format(ws, 'H3', data.get('supporter'))          # 担当支援員
        set_cell_value_preserve_format(ws, 'P3', data.get('interviewDate'))      # 面談実施日
        
        set_cell_value_preserve_format(ws, 'D4', data.get('guardianName'))       # 保護者氏名
        set_cell_value_preserve_format(ws, 'J4', data.get('childName'))          # 児童氏名
        set_cell_value_preserve_format(ws, 'P4', data.get('gender'))             # 性別
        
        set_cell_value_preserve_format(ws, 'D5', data.get('schoolName'))         # 学校名
        set_cell_value_preserve_format(ws, 'J5', str(data.get('grade', '')))     # 学年
        set_cell_value_preserve_format(ws, 'O5', data.get('singleParent'))       # ひとり親世帯
        
        # === 課題チェックリスト ===
        if 'issues' in data:
            issues_text = format_issues_text(data['issues'])
            set_cell_value_preserve_format(ws, 'B11', issues_text)
        
        # === 希望する進路 ===
        if 'futurePath' in data:
            future_path = data['futurePath']
            checkbox_進学 = '■' if future_path.get('type') == '進学' else '□'
            checkbox_就職 = '■' if future_path.get('type') == '就職' else '□'
            
            future_path_text = (
                f"確認日　{data.get('confirmDate', '')}\n"
                f"{checkbox_進学}進学　　{checkbox_就職}就職\n"
                f"（具体的内容）\n"
                f"・{future_path.get('detail', '')}"
            )
            set_cell_value_preserve_format(ws, 'B18', future_path_text)
        
        # === 短期目標（Row 29-30）===
        if 'shortTermPlan' in data:
            plan = data['shortTermPlan']
            set_cell_value_preserve_format(ws, 'B29', plan.get('issue'))           # 課題
            set_cell_value_preserve_format(ws, 'C29', plan.get('currentStatus'))   # 現状
            set_cell_value_preserve_format(ws, 'G29', plan.get('needsChild'))      # ニーズ本人
            set_cell_value_preserve_format(ws, 'G30', plan.get('needsGuardian'))   # ニーズ保護者
            set_cell_value_preserve_format(ws, 'J29', plan.get('goal'))            # 目標
            set_cell_value_preserve_format(ws, 'N29', plan.get('method'))          # 具体的な方法
        
        # === 長期目標（Row 35-36）===
        if 'longTermPlan' in data:
            plan = data['longTermPlan']
            set_cell_value_preserve_format(ws, 'B35', plan.get('issue'))           # 課題
            set_cell_value_preserve_format(ws, 'C35', plan.get('currentStatus'))   # 現状
            set_cell_value_preserve_format(ws, 'G35', plan.get('needsChild'))      # ニーズ本人
            set_cell_value_preserve_format(ws, 'G36', plan.get('needsGuardian'))   # ニーズ保護者
            set_cell_value_preserve_format(ws, 'J35', plan.get('goal'))            # 目標
            set_cell_value_preserve_format(ws, 'N35', plan.get('method'))          # 具体的な方法
        
        print('✓ データ書き込み完了')
        print()
        
        # 一時ファイルに保存
        temp_path = output_path + '.tmp'
        wb.save(temp_path)
        print(f'✓ 一時ファイルを保存: {temp_path}')
        
        # パスワード保護
        if password:
            print(f'🔒 パスワード保護を適用中...')
            add_password_protection(temp_path, output_path, password)
            os.remove(temp_path)
            print(f'✓ パスワード保護完了')
        else:
            shutil.move(temp_path, output_path)
            print('⚠️  パスワード保護なし')
        
        print()
        print(f'✅ アセスメントシートを作成: {output_path}')
        
        return {
            'success': True,
            'outputPath': output_path
        }
        
    except Exception as e:
        print(f'❌ エラー: {e}')
        import traceback
        traceback.print_exc()
        raise


def add_password_protection(input_path, output_path, password):
    """
    Excelファイルにパスワード保護を追加
    
    Args:
        input_path: 入力ファイルパス
        output_path: 出力ファイルパス
        password: パスワード
    """
    try:
        import msoffcrypto
        import io
        
        # ファイルを暗号化
        with open(input_path, 'rb') as input_file:
            ms_file = msoffcrypto.OfficeFile(input_file)
            ms_file.load_key(password=password)
            
            with open(output_path, 'wb') as output_file:
                ms_file.encrypt(password, output_file)
    except ImportError:
        # msoffcryptoがインストールされていない場合は、パスワード保護なしで保存
        print('⚠️  msoffcryptoがインストールされていません。パスワード保護なしで保存します。')
        shutil.copy2(input_path, output_path)


# テスト実行
if __name__ == '__main__':
    import sys
    import json
    
    if len(sys.argv) > 1:
        # コマンドライン引数からデータを読み込み
        data_json_path = sys.argv[1]
        with open(data_json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        template_path = data.get('templatePath', 'templates/アセスメントシート原本.xlsx')
        output_path = data.get('outputPath', 'output.xlsx')
        password = data.get('password')
        
    else:
        # テストデータ
        template_path = 'templates/アセスメントシート原本.xlsx'
        output_path = 'test_assessment.xlsx'
        password = None
        
        data = {
            'supportNumber': 'TEST-001',
            'supporter': '田中支援員',
            'interviewDate': '5月15日',
            'guardianName': '山田花子',
            'childName': '山田太郎',
            'gender': '男性',
            'schoolName': '登美丘中学校',
            'grade': 2,
            'singleParent': '該当しない',
            'confirmDate': '2025年5月15日',
            'issues': {
                '不登校': {'checked': True, 'detail': '週0回'},
                '生活リズム': {'checked': True, 'detail': '昼夜逆転'},
                '学習の遅れ・低学力': {'checked': True, 'detail': '小学生レベル'},
                '対人緊張の高さ': {'checked': True, 'detail': '初対面で緊張'},
            },
            'futurePath': {
                'type': '進学',
                'detail': '通信制高校を希望'
            },
            'shortTermPlan': {
                'issue': '生活リズムの改善',
                'currentStatus': '昼夜逆転、起床11時頃',
                'needsChild': '朝起きられるようになりたい',
                'needsGuardian': '規則正しい生活を送ってほしい',
                'goal': '9時までに起床できるようになる',
                'method': '段階的に起床時間を早める。アラーム設定、朝のルーティン作成'
            },
            'longTermPlan': {
                'issue': '進学準備',
                'currentStatus': '学習習慣なし、基礎学力不足',
                'needsChild': '高校に進学したい',
                'needsGuardian': '高校卒業まで支援してほしい',
                'goal': '通信制高校に合格する',
                'method': '基礎学力の補習、学習習慣の定着、進路相談'
            }
        }
    
    result = generate_assessment_sheet(template_path, output_path, data, password)
    
    if result['success']:
        print()
        print('=' * 70)
        print('完了!')
        print('=' * 70)
        print(json.dumps(result, ensure_ascii=False, indent=2))
